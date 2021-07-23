#! python3
# -- coding: utf-8 --

import openpyxl
# import logging
# logging.disable(logging.CRITICAL)
# logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')
# logging.debug("Start of program.")


def blankRowInserter(excelfile, n, m):
    wb = openpyxl.load_workbook(excelfile)
    ws = wb.active
    for i in sorted(range(n + 1, ws.max_row +1), reverse=True):
        for c in ws[i]:
            ws[str(c.column) + str(c.row + m)] = c.value
    for i in range(n + 1, n + m +1):
        for c in ws[i]:
            c.value = None
    wb.save(excelfile)
    wb.close()
    print('-----Done-----')


# blankRowInserter('Multiplication.xlsx', 2, 3)
