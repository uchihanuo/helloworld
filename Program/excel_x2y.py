#! python3
# -- coding: utf-8 --

import openpyxl
from openpyxl.utils import get_column_letter
# import logging
# logging.disable(logging.CRITICAL)
# logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')
# logging.debug("Start of program.")


def excel_x2y(excelfile):
    wb = openpyxl.load_workbook(excelfile)
    ws_previous = wb.active
    ws = wb.create_sheet(title=ws_previous.title + '_x2y')
    rows = []
    for i in range(1, ws_previous.max_row + 1):
        a_list = []
        for row in ws_previous[i]:
            a_list += [row.value]
        rows.append(a_list)
    for i in range(ws_previous.max_row):
        for j in range(ws_previous.max_column):
            ws[get_column_letter(i + 1) + str(j + 1)] = rows[i][j]
    wb.save(excelfile)
    wb.close()


excel_x2y('Multiplication.xlsx')
